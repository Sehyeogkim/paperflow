"""Safe, bounded extraction of project artifacts into a portable manifest.

The public entry point, :func:`build_artifact_manifest`, accepts a project root that
has already been selected by the server.  It never accepts an arbitrary output path,
never serializes the absolute project root, and refuses files that resolve outside the
root.  By default only ``data/`` and ``reference/`` are scanned.

Extraction is deliberately shallow: it creates small, provenance-addressable text
chunks for downstream LLM use.  It is not a general document conversion service.
Every parser has byte/row/page/character limits and a bad artifact is represented by a
status on that artifact rather than aborting the whole manifest.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import mimetypes
import re
import unicodedata
import zipfile
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union


_SUPPORTED = {
    ".txt": "text",
    ".dat": "text",
    ".md": "markdown",
    ".csv": "csv",
    ".json": "json",
    ".pdf": "pdf",
    ".xlsx": "xlsx",
}

_MIME_BY_KIND = {
    "text": "text/plain",
    "markdown": "text/markdown",
    "csv": "text/csv",
    "json": "application/json",
    "pdf": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


@dataclass(frozen=True)
class ExtractionLimits:
    """Hard limits for one manifest build.

    ``max_file_bytes`` limits parsing, while the SHA-256 is streamed so the manifest
    can still identify an over-limit file.  ``max_total_bytes`` stops hashing/parsing
    once a whole-project budget is exhausted.
    """

    max_files: int = 100
    max_total_bytes: int = 50 * 1024 * 1024
    max_file_bytes: int = 10 * 1024 * 1024
    max_chars_per_artifact: int = 50_000
    max_chunk_chars: int = 4_000
    max_text_lines: int = 2_000
    max_csv_rows: int = 200
    max_csv_columns: int = 100
    max_cell_chars: int = 1_000
    max_pdf_pages: int = 25
    max_pdf_chars_per_page: int = 10_000
    max_xlsx_sheets: int = 10
    max_xlsx_rows_per_sheet: int = 50
    max_xlsx_columns: int = 50
    max_xlsx_members: int = 2_000
    max_xlsx_uncompressed_bytes: int = 25 * 1024 * 1024

    def __post_init__(self) -> None:
        for name, value in asdict(self).items():
            if not isinstance(value, int) or value <= 0:
                raise ValueError("%s must be a positive integer" % name)


@dataclass
class ProvenanceLocator:
    """A locator that lets a consumer trace a chunk back to its source."""

    relative_path: str
    locator_type: str
    line_start: Optional[int] = None
    line_end: Optional[int] = None
    row_start: Optional[int] = None
    row_end: Optional[int] = None
    page: Optional[int] = None
    sheet: Optional[str] = None
    char_start: Optional[int] = None
    char_end: Optional[int] = None
    json_pointer: Optional[str] = None


@dataclass
class ArtifactChunk:
    chunk_id: str
    text: str
    locator: ProvenanceLocator


@dataclass
class ArtifactRecord:
    logical_id: str
    relative_path: str
    original_name: str
    sha256: Optional[str]
    size_bytes: Optional[int]
    mime_type: str
    extension: str
    classification: str
    status: str
    chunks: List[ArtifactChunk] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    error_code: Optional[str] = None
    error_message: Optional[str] = None


@dataclass
class ArtifactManifest:
    """Serializable output of a bounded project artifact scan."""

    manifest_version: str = "1.0"
    artifacts: List[ArtifactRecord] = field(default_factory=list)
    scanned_roots: List[str] = field(default_factory=list)
    truncated: bool = False
    warnings: List[str] = field(default_factory=list)
    limits: Dict[str, int] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    def to_json(self, *, indent: Optional[int] = 2) -> str:
        return json.dumps(self.to_dict(), ensure_ascii=False, indent=indent)


def _safe_error(exc: BaseException) -> str:
    """Return a short error without paths, payloads, or provider details."""

    if isinstance(exc, OSError):
        return exc.__class__.__name__
    message = str(exc).replace("\n", " ").replace("\r", " ")
    return (message[:197] + "...") if len(message) > 200 else message


def _logical_id(relative_path: str) -> str:
    normalized = unicodedata.normalize("NFKD", relative_path)
    ascii_name = normalized.encode("ascii", "ignore").decode("ascii").lower()
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_name).strip("-") or "artifact"
    digest = hashlib.sha256(relative_path.encode("utf-8")).hexdigest()[:10]
    return "%s-%s" % (slug[:48].rstrip("-"), digest)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _is_within(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


def _contains_symlink(path: Path, root: Path) -> bool:
    """Return true when any path component below root is a symlink."""

    try:
        relative = path.relative_to(root)
    except ValueError:
        return True
    current = root
    for part in relative.parts:
        current = current / part
        if current.is_symlink():
            return True
    return False


def _under_scan_root(path: Path, scan_roots: Sequence[str]) -> bool:
    for scan_root in scan_roots:
        try:
            path.relative_to(Path(scan_root))
            return True
        except ValueError:
            continue
    return False


def _mime_type(path: Path, classification: str) -> str:
    if classification in _MIME_BY_KIND:
        return _MIME_BY_KIND[classification]
    guessed, _ = mimetypes.guess_type(path.name)
    return guessed or "application/octet-stream"


def _base_record(relative_path: str, *, path: Optional[Path] = None) -> ArtifactRecord:
    extension = Path(relative_path).suffix.lower()
    classification = _SUPPORTED.get(extension, "unsupported")
    return ArtifactRecord(
        logical_id=_logical_id(relative_path),
        relative_path=relative_path,
        original_name=Path(relative_path).name,
        sha256=None,
        size_bytes=None,
        mime_type=_mime_type(path or Path(relative_path), classification),
        extension=extension,
        classification=classification,
        status="pending",
    )


def _chunk_text(
    record: ArtifactRecord,
    text: str,
    locator_type: str,
    limits: ExtractionLimits,
    *,
    line_offset: int = 1,
) -> List[ArtifactChunk]:
    """Chunk text on line boundaries and preserve 1-based source line numbers."""

    lines = text.splitlines()
    chunks: List[ArtifactChunk] = []
    current: List[str] = []
    current_chars = 0
    start_line = line_offset

    def flush(end_line: int) -> None:
        nonlocal current, current_chars, start_line
        if not current:
            return
        body = "\n".join(current)
        chunks.append(ArtifactChunk(
            chunk_id="%s:c%04d" % (record.logical_id, len(chunks) + 1),
            text=body,
            locator=ProvenanceLocator(
                relative_path=record.relative_path,
                locator_type=locator_type,
                line_start=start_line,
                line_end=end_line,
            ),
        ))
        current = []
        current_chars = 0
        start_line = end_line + 1

    for index, line in enumerate(lines, start=line_offset):
        # One pathological line must not defeat the chunk character limit.
        if len(line) > limits.max_chunk_chars:
            flush(index - 1)
            for offset in range(0, len(line), limits.max_chunk_chars):
                piece = line[offset:offset + limits.max_chunk_chars]
                chunks.append(ArtifactChunk(
                    chunk_id="%s:c%04d" % (record.logical_id, len(chunks) + 1),
                    text=piece,
                    locator=ProvenanceLocator(
                        relative_path=record.relative_path,
                        locator_type=locator_type,
                        line_start=index,
                        line_end=index,
                        char_start=offset,
                        char_end=offset + len(piece),
                    ),
                ))
            start_line = index + 1
            continue
        extra = len(line) + (1 if current else 0)
        if current and current_chars + extra > limits.max_chunk_chars:
            flush(index - 1)
            start_line = index
        current.append(line)
        current_chars += extra
    flush(line_offset + len(lines) - 1)
    return chunks


def _extract_text(path: Path, record: ArtifactRecord, limits: ExtractionLimits) -> None:
    raw = path.read_bytes()
    text = raw.decode("utf-8", errors="replace")
    lines = text.splitlines()[:limits.max_text_lines]
    text = "\n".join(lines)[:limits.max_chars_per_artifact]
    record.chunks = _chunk_text(record, text, "text_lines", limits)
    record.metadata = {
        "encoding": "utf-8",
        "replacement_characters": text.count("\ufffd"),
        "lines_extracted": len(text.splitlines()),
        "truncated": len(raw) > len(text.encode("utf-8", errors="replace")),
    }


def _clean_cell(value: Any, limit: int) -> str:
    if value is None:
        return ""
    text = str(value).replace("\x00", "")
    return text[:limit]


def _trim_empty_tail(row: Sequence[str]) -> List[str]:
    values = list(row)
    while values and values[-1] == "":
        values.pop()
    return values


def _rows_as_delimited(rows: Sequence[Sequence[str]], delimiter: str) -> str:
    out = io.StringIO()
    writer = csv.writer(out, delimiter=delimiter, lineterminator="\n")
    writer.writerows(rows)
    return out.getvalue().rstrip("\n")


def _append_table_chunks(
    record: ArtifactRecord,
    rows: Sequence[Sequence[str]],
    limits: ExtractionLimits,
    *,
    delimiter: str,
    locator_type: str,
    sheet: Optional[str] = None,
    char_budget: Optional[int] = None,
) -> int:
    """Append bounded row chunks and return the number of characters emitted."""

    budget = limits.max_chars_per_artifact if char_budget is None else max(0, char_budget)
    emitted = 0
    pending: List[Sequence[str]] = []
    pending_start = 1

    def append_piece(text: str, row_start: int, row_end: int,
                     char_start: Optional[int] = None) -> None:
        nonlocal budget, emitted
        if not text or budget <= 0:
            return
        piece = text[:min(limits.max_chunk_chars, budget)]
        record.chunks.append(ArtifactChunk(
            chunk_id="%s:c%04d" % (record.logical_id, len(record.chunks) + 1),
            text=piece,
            locator=ProvenanceLocator(
                relative_path=record.relative_path,
                locator_type=locator_type,
                row_start=row_start,
                row_end=row_end,
                sheet=sheet,
                char_start=char_start,
                char_end=(char_start + len(piece)) if char_start is not None else None,
            ),
        ))
        emitted += len(piece)
        budget -= len(piece)

    def flush_pending(row_end: int) -> None:
        nonlocal pending, pending_start
        if pending:
            append_piece(_rows_as_delimited(pending, delimiter), pending_start, row_end)
            pending = []
        pending_start = row_end + 1

    for row_number, row in enumerate(rows, start=1):
        if budget <= 0:
            break
        rendered = _rows_as_delimited([row], delimiter)
        if not rendered:
            continue
        if len(rendered) > limits.max_chunk_chars:
            flush_pending(row_number - 1)
            for offset in range(0, len(rendered), limits.max_chunk_chars):
                append_piece(rendered[offset:offset + limits.max_chunk_chars],
                             row_number, row_number, offset)
            pending_start = row_number + 1
            continue
        candidate = _rows_as_delimited(pending + [row], delimiter)
        if pending and len(candidate) > limits.max_chunk_chars:
            flush_pending(row_number - 1)
            pending_start = row_number
        pending.append(row)
    flush_pending(pending_start + len(pending) - 1)
    return emitted


def _extract_csv(path: Path, record: ArtifactRecord, limits: ExtractionLimits) -> None:
    text = path.read_bytes().decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows: List[List[str]] = []
    truncated = False
    for index, row in enumerate(reader):
        if index >= limits.max_csv_rows:
            truncated = True
            break
        rows.append([_clean_cell(v, limits.max_cell_chars)
                     for v in row[:limits.max_csv_columns]])

    header = rows[0] if rows else []
    record.metadata = {
        "header": header,
        "sample_rows": rows[1:11],
        "rows_extracted": len(rows),
        "columns_extracted": max((len(r) for r in rows), default=0),
        "truncated": truncated,
    }
    emitted = _append_table_chunks(
        record, rows, limits, delimiter=",", locator_type="csv_rows")
    if emitted >= limits.max_chars_per_artifact:
        record.metadata["truncated"] = True


def _extract_json(path: Path, record: ArtifactRecord, limits: ExtractionLimits) -> None:
    raw = path.read_bytes().decode("utf-8", errors="strict")
    value = json.loads(raw)
    rendered = json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True)
    truncated = len(rendered) > limits.max_chars_per_artifact
    rendered = rendered[:limits.max_chars_per_artifact]
    record.metadata = {
        "top_level_type": type(value).__name__,
        "top_level_keys": list(value.keys())[:100] if isinstance(value, dict) else [],
        "truncated": truncated,
    }
    for offset in range(0, len(rendered), limits.max_chunk_chars):
        text = rendered[offset:offset + limits.max_chunk_chars]
        record.chunks.append(ArtifactChunk(
            chunk_id="%s:c%04d" % (record.logical_id, len(record.chunks) + 1),
            text=text,
            locator=ProvenanceLocator(
                relative_path=record.relative_path,
                locator_type="json_text",
                char_start=offset,
                char_end=offset + len(text),
                json_pointer="$",
            ),
        ))


def _extract_pdf(path: Path, record: ArtifactRecord, limits: ExtractionLimits) -> None:
    with path.open("rb") as stream:
        signature = stream.read(5)
    if signature != b"%PDF-":
        raise ValueError("invalid PDF signature")
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - packaging failure, exercised by integration
        raise RuntimeError("pypdf is not installed") from exc

    reader = PdfReader(str(path), strict=False)
    if reader.is_encrypted:
        try:
            unlocked = reader.decrypt("")
        except Exception as exc:
            raise ValueError("encrypted PDF is not supported") from exc
        if not unlocked:
            raise ValueError("encrypted PDF is not supported")
    page_count = len(reader.pages)
    total_chars = 0
    pages_extracted = 0
    for page_number, page in enumerate(reader.pages[:limits.max_pdf_pages], start=1):
        remaining = limits.max_chars_per_artifact - total_chars
        if remaining <= 0:
            break
        text = (page.extract_text() or "")[:limits.max_pdf_chars_per_page]
        text = text[:remaining]
        for offset in range(0, len(text), limits.max_chunk_chars):
            piece = text[offset:offset + limits.max_chunk_chars]
            record.chunks.append(ArtifactChunk(
                chunk_id="%s:c%04d" % (record.logical_id, len(record.chunks) + 1),
                text=piece,
                locator=ProvenanceLocator(
                    relative_path=record.relative_path,
                    locator_type="pdf_page",
                    page=page_number,
                    char_start=offset,
                    char_end=offset + len(piece),
                ),
            ))
        total_chars += len(text)
        pages_extracted += 1
    record.metadata = {
        "pages_total": page_count,
        "pages_extracted": pages_extracted,
        "truncated": page_count > pages_extracted,
    }


def _xlsx_preflight(path: Path, limits: ExtractionLimits) -> None:
    with path.open("rb") as stream:
        signature = stream.read(2)
    if signature != b"PK":
        raise ValueError("invalid XLSX signature")
    with zipfile.ZipFile(path) as archive:
        members = archive.infolist()
        if len(members) > limits.max_xlsx_members:
            raise ValueError("XLSX member limit exceeded")
        if sum(member.file_size for member in members) > limits.max_xlsx_uncompressed_bytes:
            raise ValueError("XLSX uncompressed byte limit exceeded")


def _extract_xlsx(path: Path, record: ArtifactRecord, limits: ExtractionLimits) -> None:
    _xlsx_preflight(path, limits)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - packaging failure, exercised by integration
        raise RuntimeError("openpyxl is not installed") from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True, keep_links=False)
    sheet_summaries: List[Dict[str, Any]] = []
    total_chars = 0
    sheet_count = len(workbook.sheetnames)
    try:
        worksheets = workbook.worksheets[:limits.max_xlsx_sheets]
        for worksheet in worksheets:
            rows: List[List[str]] = []
            truncated = False
            for row_index, row in enumerate(
                worksheet.iter_rows(max_col=limits.max_xlsx_columns, values_only=True),
                start=1,
            ):
                if row_index > limits.max_xlsx_rows_per_sheet:
                    truncated = True
                    break
                rows.append(_trim_empty_tail(
                    [_clean_cell(value, limits.max_cell_chars) for value in row]))
            header = rows[0] if rows else []
            summary = {
                "sheet": worksheet.title,
                "header": header,
                "sample_rows": rows[1:11],
                "rows_extracted": len(rows),
                "columns_extracted": max((len(row) for row in rows), default=0),
                "truncated": truncated,
            }
            sheet_summaries.append(summary)
            if not rows or total_chars >= limits.max_chars_per_artifact:
                continue
            total_chars += _append_table_chunks(
                record,
                rows,
                limits,
                delimiter="\t",
                locator_type="xlsx_sheet_rows",
                sheet=worksheet.title,
                char_budget=limits.max_chars_per_artifact - total_chars,
            )
    finally:
        workbook.close()
    record.metadata = {
        "sheets_total": sheet_count,
        "sheets_extracted": len(sheet_summaries),
        "sheets": sheet_summaries,
        "truncated": sheet_count > len(sheet_summaries)
                     or total_chars >= limits.max_chars_per_artifact,
    }


_EXTRACTORS = {
    "text": _extract_text,
    "markdown": _extract_text,
    "csv": _extract_csv,
    "json": _extract_json,
    "pdf": _extract_pdf,
    "xlsx": _extract_xlsx,
}


def _blocked_record(relative_path: str, code: str, message: str) -> ArtifactRecord:
    record = _base_record(relative_path)
    record.status = "blocked"
    record.error_code = code
    record.error_message = message
    return record


def _candidate_paths(
    root: Path,
    relative_paths: Optional[Iterable[str]],
    scan_roots: Sequence[str],
) -> Tuple[List[Tuple[str, Path]], List[ArtifactRecord]]:
    candidates: List[Tuple[str, Path]] = []
    blocked: List[ArtifactRecord] = []
    seen = set()
    if relative_paths is not None:
        for supplied in relative_paths:
            supplied_text = str(supplied).replace("\\", "/")
            supplied_path = Path(supplied_text)
            if supplied_path.is_absolute():
                safe_name = Path(supplied_text).name or "artifact"
                blocked.append(_blocked_record(
                    "rejected/%s" % safe_name,
                    "absolute_path",
                    "absolute artifact paths are not allowed",
                ))
                continue
            relative = supplied_path.as_posix()
            if ".." in supplied_path.parts:
                blocked.append(_blocked_record(
                    relative, "path_escape", "parent traversal is not allowed"))
                continue
            if any(part.startswith(".") for part in supplied_path.parts):
                blocked.append(_blocked_record(
                    relative, "hidden_path", "hidden artifact paths are not allowed"))
                continue
            if not _under_scan_root(supplied_path, scan_roots):
                blocked.append(_blocked_record(
                    relative, "outside_scan_roots",
                    "artifact is outside the configured scan roots"))
                continue
            if relative in seen:
                continue
            seen.add(relative)
            candidates.append((relative, root / supplied_path))
        return candidates, blocked

    for scan_root in scan_roots:
        scan_path = root / scan_root
        if not scan_path.is_dir():
            continue
        for path in sorted(scan_path.rglob("*")):
            relative_path = path.relative_to(root)
            if any(part.startswith(".") for part in relative_path.parts):
                continue
            if path.is_file() or path.is_symlink():
                relative = relative_path.as_posix()
                if relative not in seen:
                    seen.add(relative)
                    candidates.append((relative, path))
    return candidates, blocked


def build_artifact_manifest(
    project_root: Union[Path, str],
    *,
    relative_paths: Optional[Iterable[str]] = None,
    scan_roots: Sequence[str] = ("data", "reference"),
    limits: Optional[ExtractionLimits] = None,
) -> ArtifactManifest:
    """Build a bounded manifest for artifacts below a server-selected project root.

    ``relative_paths`` is optional and, when supplied, must contain only paths relative
    to ``project_root``.  The output contains relative provenance only.  Unsupported,
    malformed, too-large, and unsafe files are records with a non-``extracted`` status;
    they do not abort extraction of other files.
    """

    limits = limits or ExtractionLimits()
    root = Path(project_root).resolve(strict=True)
    if not root.is_dir():
        raise NotADirectoryError("project_root must be a directory")
    safe_scan_roots = []
    for item in scan_roots:
        p = Path(item)
        if p.is_absolute() or ".." in p.parts:
            raise ValueError("scan_roots must be relative directories without '..'")
        safe_scan_roots.append(p.as_posix())

    if isinstance(relative_paths, (str, Path)):
        relative_paths = [str(relative_paths)]
    candidates, blocked = _candidate_paths(root, relative_paths, safe_scan_roots)
    manifest = ArtifactManifest(
        artifacts=blocked,
        scanned_roots=safe_scan_roots if relative_paths is None else [],
        limits=asdict(limits),
    )
    total_bytes = 0
    for relative, candidate in candidates:
        if len(manifest.artifacts) >= limits.max_files:
            manifest.truncated = True
            manifest.warnings.append("artifact count limit reached")
            break
        record = _base_record(relative, path=candidate)
        manifest.artifacts.append(record)
        if _contains_symlink(candidate, root):
            record.status = "blocked"
            record.error_code = "symlink_not_allowed"
            record.error_message = "symlink artifacts are not allowed"
            continue
        try:
            resolved = candidate.resolve(strict=True)
        except (FileNotFoundError, RuntimeError) as exc:
            record.status = "error"
            record.error_code = "not_found"
            record.error_message = _safe_error(exc)
            continue
        if not _is_within(resolved, root):
            record.status = "blocked"
            record.error_code = "path_escape"
            record.error_message = "artifact resolves outside project root"
            continue
        if not resolved.is_file():
            record.status = "unsupported"
            record.error_code = "not_regular_file"
            continue
        try:
            record.size_bytes = resolved.stat().st_size
        except OSError as exc:
            record.status = "error"
            record.error_code = "stat_error"
            record.error_message = _safe_error(exc)
            continue
        if total_bytes + record.size_bytes > limits.max_total_bytes:
            record.status = "limit_exceeded"
            record.error_code = "project_byte_limit"
            record.error_message = "project artifact byte limit exceeded"
            manifest.truncated = True
            if "project artifact byte limit reached" not in manifest.warnings:
                manifest.warnings.append("project artifact byte limit reached")
            continue
        total_bytes += record.size_bytes
        try:
            record.sha256 = _sha256(resolved)
        except OSError as exc:
            record.status = "error"
            record.error_code = "hash_error"
            record.error_message = _safe_error(exc)
            continue
        if record.classification == "unsupported":
            record.status = "unsupported"
            record.error_code = "unsupported_extension"
            continue
        if record.size_bytes > limits.max_file_bytes:
            record.status = "limit_exceeded"
            record.error_code = "file_byte_limit"
            record.error_message = "artifact exceeds extraction byte limit"
            continue
        extractor = _EXTRACTORS[record.classification]
        try:
            extractor(resolved, record, limits)
            record.status = "extracted"
        except (OSError, UnicodeError, ValueError, TypeError, KeyError, csv.Error,
                zipfile.BadZipFile, RuntimeError) as exc:
            record.status = "error"
            record.error_code = "%s_error" % record.classification
            record.error_message = _safe_error(exc)
            record.chunks = []
    return manifest


def extract_project_artifacts(
    project_root: Union[Path, str],
    *,
    limits: Optional[ExtractionLimits] = None,
) -> ArtifactManifest:
    """Convenience alias for the default ``data/`` + ``reference/`` scan."""

    return build_artifact_manifest(project_root, limits=limits)
